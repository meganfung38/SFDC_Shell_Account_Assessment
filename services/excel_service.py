from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import io
import pandas as pd

class ExcelService:
    """Service for Excel operations and Account data handling"""
    
    def __init__(self):
        # RingCentral Brand Colors
        self.rc_cerulean = "0684BC"      # RingCentral primary blue
        self.rc_orange = "FF7A00"        # RingCentral orange
        self.rc_ocean = "002855"         # RingCentral dark blue
        self.rc_linen = "F1EFEC"         # RingCentral background
        self.rc_ash = "C8C2B4"           # RingCentral light gray
        self.rc_warm_black = "2B2926"    # RingCentral dark gray
        
        # Excel Styling with RingCentral Colors
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color=self.rc_cerulean, end_color=self.rc_cerulean, fill_type="solid")
        self.summary_font = Font(bold=True, color=self.rc_ocean)
        self.summary_fill = PatternFill(start_color=self.rc_linen, end_color=self.rc_linen, fill_type="solid")
        self.title_font = Font(bold=True, size=16, color=self.rc_ocean)
        self.accent_fill = PatternFill(start_color=self.rc_orange, end_color=self.rc_orange, fill_type="solid")
        self.border = Border(
            left=Side(style='thin', color=self.rc_ash),
            right=Side(style='thin', color=self.rc_ash),
            top=Side(style='thin', color=self.rc_ash),
            bottom=Side(style='thin', color=self.rc_ash)
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    

    
    
    

    
    def parse_excel_file(self, file_content):
        """Parse uploaded Excel file and return sheet names and preview data"""
        try:
            # Load workbook from file content
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            sheet_names = wb.sheetnames
            
            # Get preview data from first sheet
            first_sheet = wb[sheet_names[0]]
            
            # Read first 10 rows for preview
            preview_data = []
            headers = []
            
            for row_idx, row in enumerate(first_sheet.iter_rows(values_only=True)):
                if row_idx == 0:
                    # First row as headers
                    headers = [cell if cell is not None else f"Column_{i+1}" for i, cell in enumerate(row)]
                elif row_idx < 11:  # First 10 data rows
                    row_data = [cell if cell is not None else "" for cell in row]
                    # Pad row to match header length
                    while len(row_data) < len(headers):
                        row_data.append("")
                    preview_data.append(row_data[:len(headers)])  # Trim to header length
                else:
                    break
            
            wb.close()
            
            # Calculate total rows safely (max_row can be None for empty sheets)
            total_rows = (first_sheet.max_row - 1) if first_sheet.max_row else 0
            
            return {
                'success': True,
                'sheet_names': sheet_names,
                'headers': headers,
                'preview_data': preview_data,
                'total_rows': max(total_rows, len(preview_data))  # Use actual data count if max_row is unreliable
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error parsing Excel file: {str(e)}"
            }
    


    def extract_account_ids_from_excel(self, file_content, sheet_name, account_id_column):
        """Extract Account IDs from specified column in Excel file"""
        try:
            # Use pandas for easier data extraction - read as string to preserve Account ID format
            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name, dtype={account_id_column: str})
            
            if account_id_column not in df.columns:
                return {
                    'success': False,
                    'error': f"Column '{account_id_column}' not found in sheet '{sheet_name}'"
                }
            
            # Extract Account IDs and remove null/empty values
            account_ids = df[account_id_column].dropna().astype(str).tolist()
            # Remove empty strings and whitespace-only strings, and handle Excel formatting issues
            cleaned_account_ids = []
            for aid in account_ids:
                aid_str = str(aid).strip()
                # Remove any Excel formatting artifacts
                if aid_str and aid_str.lower() not in ['nan', 'none', 'null']:
                    # Handle potential floating point conversion (e.g., "1.23456789012345e+17")
                    if 'e+' in aid_str.lower():
                        try:
                            # Convert scientific notation back to full number
                            aid_str = f"{float(aid_str):.0f}"
                        except:
                            pass
                    cleaned_account_ids.append(aid_str)
            
            account_ids = cleaned_account_ids
            
            # Get original data for later merging - handle NaN values
            # Replace NaN with empty string to avoid JSON serialization issues
            df_clean = df.where(pd.notnull(df), '')  # Replace NaN with empty string
            original_data = df_clean.to_dict('records')
            
            return {
                'success': True,
                'account_ids': account_ids,
                'original_data': original_data,
                'total_rows': len(df)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error extracting Account IDs: {str(e)}"
            }
    
    def _convert_15_to_18_char_id(self, id_15):
        """Convert 15-character Salesforce ID to 18-character format"""
        if len(id_15) != 15:
            return id_15
        
        # Salesforce ID conversion algorithm
        suffix = ""
        for i in range(3):
            chunk = id_15[i*5:(i+1)*5]
            chunk_value = 0
            for j, char in enumerate(chunk):
                if char.isupper():
                    chunk_value += 2 ** j
            
            # Convert to base-32 character
            if chunk_value < 26:
                suffix += chr(ord('A') + chunk_value)
            else:
                suffix += str(chunk_value - 26)
        
        return id_15 + suffix

    def create_basic_excel(self, data, headers, title="Data Export", filename_prefix="export"):
        """Create a basic Excel file with data and headers"""
        try:
            wb = Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Data"
                
                # Add title
                current_row = 1
                ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
                ws[f'A{current_row}'] = title
                ws[f'A{current_row}'].font = self.title_font
                ws[f'A{current_row}'].alignment = self.center_alignment
                current_row += 1
                
                # Add timestamp
                ws.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
                ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws[f'A{current_row}'].alignment = self.center_alignment
                current_row += 2
                
                # Add headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                    cell.border = self.border
                
                current_row += 1
                
                # Add data rows
                for row_data in data:
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.border = self.border
                    current_row += 1
                
                # Auto-adjust column widths
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Create file buffer
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': file_buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating Excel file: {str(e)}"
            } 

    def create_analysis_export(self, accounts, summary, export_type="analysis"):
        """Create Excel export for analysis results with single table format"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Account Analysis"
            
            # Add title to sheet
            current_row = 1
            title = f"SFDC Account Analysis - {export_type.replace('_', ' ').title()}"
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = title
            ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
            ws[f'A{current_row}'].alignment = self.center_alignment
            ws[f'A{current_row}'].fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
            current_row += 1
            
            # Add timestamp
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws[f'A{current_row}'].alignment = self.center_alignment
            current_row += 2
            
            # Add Summary Table
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "📊 Analysis Summary"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C5AA0")
            ws[f'A{current_row}'].fill = PatternFill(start_color="F1EFEC", end_color="F1EFEC", fill_type="solid")
            current_row += 1
            
            # Summary table headers
            summary_headers = ["Metric", "Value"]
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = PatternFill(start_color="0684BC", end_color="0684BC", fill_type="solid")
                cell.alignment = self.center_alignment
                cell.border = self.border
            current_row += 1
            
            # Summary metrics
            summary_metrics = [
                ["Total Accounts Requested", summary.get('total_requested', 0)],
                ["Accounts Retrieved", summary.get('accounts_retrieved', 0)],
                ["Accounts with Shell", sum(1 for acc in accounts if acc.get('Has_Shell', False))],
                ["Average Customer Consistency Score", 
                 round(sum(acc.get('Customer_Consistency', {}).get('score', 0) for acc in accounts) / len(accounts), 1) if accounts else 0],
                ["Average Customer-Shell Coherence Score", 
                 round(sum(acc.get('Customer_Shell_Coherence', {}).get('score', 0) for acc in accounts if acc.get('Has_Shell', False)) / 
                       sum(1 for acc in accounts if acc.get('Has_Shell', False)), 1) if any(acc.get('Has_Shell', False) for acc in accounts) else 0],
                ["Accounts with Address Consistency", 
                 sum(1 for acc in accounts if acc.get('Address_Consistency', {}).get('is_consistent', False))],
                ["Average AI Confidence Score", 
                 round(sum(acc.get('AI_Assessment', {}).get('confidence_score', 0) for acc in accounts) / len(accounts), 1) if accounts else 0]
            ]
            
            for metric in summary_metrics:
                ws[f'A{current_row}'] = metric[0]
                ws[f'B{current_row}'] = metric[1]
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].border = self.border
                ws[f'B{current_row}'].border = self.border
                current_row += 1
            
            current_row += 2
            
            # Add Analysis Table Header
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "📋 Account Analysis Results"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C5AA0")
            ws[f'A{current_row}'].fill = PatternFill(start_color="F1EFEC", end_color="F1EFEC", fill_type="solid")
            current_row += 1
            
            # Define headers for analysis table
            headers = [
                # Account Identification (Frozen)
                "Account ID", "Account Name",
                # Account Metadata
                "Record Type", "Parent ID", "Parent Name", "Website", 
                "Billing State", "Billing Country", "Billing Postal Code",
                "ZI Company", "ZI Website", "ZI State", "ZI Country", "ZI Postal Code",
                "Contact Most Frequent Email",
                # Assessment Flags
                "Bad Domain", "Bad Domain Explanation",
                "Has Shell", "Has Shell Explanation",
                "Customer Consistency Score", "Customer Consistency Explanation",
                "Customer-Shell Coherence Score", "Customer-Shell Coherence Explanation",
                "Address Consistency", "Address Consistency Explanation",
                # AI Assessment
                "AI Confidence Score", "AI Analysis"
            ]
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = PatternFill(start_color="0684BC", end_color="0684BC", fill_type="solid")
                cell.alignment = self.center_alignment
                cell.border = self.border
            current_row += 1
            
            # Add data rows
            for row_idx, account in enumerate(accounts, current_row):
                # Get flag explanations
                has_shell_explanation = "Account has parent shell relationship" if account.get('Has_Shell', False) else "No parent shell relationship"
                customer_consistency_explanation = account.get('Customer_Consistency', {}).get('explanation', '')
                customer_shell_coherence_explanation = account.get('Customer_Shell_Coherence', {}).get('explanation', '') if account.get('Has_Shell', False) else 'N/A - No shell relationship'
                address_consistency_explanation = account.get('Address_Consistency', {}).get('explanation', '')
                
                # Get AI analysis
                ai_assessment = account.get('AI_Assessment', {})
                ai_bullets = ai_assessment.get('explanation_bullets', [])
                ai_analysis = "\n".join([f"• {bullet}" for bullet in ai_bullets]) if ai_bullets else "No AI analysis available"
                
                # Get assessment flag explanations
                bad_domain_explanation = account.get('Bad_Domain', {}).get('explanation', '')
                
                # Create row data
                row_data = [
                    # Account Identification
                    account.get('Id', ''),
                    account.get('Name', ''),
                    # Account Metadata
                    account.get('RecordType', {}).get('Name', ''),
                    account.get('ParentId', ''),
                    account.get('Parent', {}).get('Name', '') if account.get('Parent') else '',
                    account.get('Website', ''),
                    account.get('BillingState', ''),
                    account.get('BillingCountry', ''),
                    account.get('BillingPostalCode', ''),
                    account.get('ZI_Company_Name__c', ''),
                    account.get('ZI_Website__c', ''),
                    account.get('ZI_Company_State__c', ''),
                    account.get('ZI_Company_Country__c', ''),
                    account.get('ZI_Company_Postal_Code__c', ''),
                    account.get('ContactMostFrequentEmail__c', ''),
                    # Assessment Flags
                    "❌ True" if account.get('Bad_Domain', {}).get('is_bad', False) else "✅ False",
                    bad_domain_explanation,
                    "✅ True" if account.get('Has_Shell', False) else "❌ False",
                    has_shell_explanation,
                    f"{account.get('Customer_Consistency', {}).get('score', 0)}/100",
                    customer_consistency_explanation,
                    f"{account.get('Customer_Shell_Coherence', {}).get('score', 0)}/100" if account.get('Has_Shell', False) else "N/A",
                    customer_shell_coherence_explanation,
                    "✅ True" if account.get('Address_Consistency', {}).get('is_consistent', False) else "❌ False",
                    address_consistency_explanation,
                    # AI Assessment
                    f"{ai_assessment.get('confidence_score', 0)}/100",
                    ai_analysis
                ]
                
                # Add row data
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = self.border
            
            # Set frozen panes (Account ID and Name columns)
            ws.freeze_panes = "C2"
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Create file buffer
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"sfdc_analysis_{export_type}_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': file_buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating analysis export: {str(e)}"
            }

    def create_excel_analysis_export(self, accounts, original_data, excel_info):
        """Create Excel export for Excel analysis with original data + AI analysis in single table"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Excel Analysis"
            
            # Add title to sheet
            current_row = 1
            title = f"Excel Analysis Results - {excel_info.get('file_name', 'Unknown File')}"
            ws.merge_cells(f'A{current_row}:Z{current_row}')
            ws[f'A{current_row}'] = title
            ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
            ws[f'A{current_row}'].alignment = self.center_alignment
            ws[f'A{current_row}'].fill = PatternFill(start_color="FF7A00", end_color="FF7A00", fill_type="solid")
            current_row += 1
            
            # Add timestamp and file info
            ws.merge_cells(f'A{current_row}:Z{current_row}')
            ws[f'A{current_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | File: {excel_info.get('file_name', 'Unknown')} | Sheet: {excel_info.get('sheet_name', 'Unknown')}"
            ws[f'A{current_row}'].alignment = self.center_alignment
            current_row += 2
            
            # Add Summary Table
            ws.merge_cells(f'A{current_row}:Z{current_row}')
            ws[f'A{current_row}'] = "📊 Analysis Summary"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C5AA0")
            ws[f'A{current_row}'].fill = PatternFill(start_color="F1EFEC", end_color="F1EFEC", fill_type="solid")
            current_row += 1
            
            # Summary table headers
            summary_headers = ["Metric", "Value"]
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = PatternFill(start_color="0684BC", end_color="0684BC", fill_type="solid")
                cell.alignment = self.center_alignment
                cell.border = self.border
            current_row += 1
            
            # Summary metrics
            def get_ai_confidence_score(account):
                ai_assessment = account.get('AI_Assessment', {})
                
                # Handle case where AI assessment is a string
                if isinstance(ai_assessment, str):
                    try:
                        import json
                        ai_assessment = json.loads(ai_assessment)
                    except (json.JSONDecodeError, TypeError):
                        return 0
                
                if isinstance(ai_assessment, dict):
                    return ai_assessment.get('confidence_score', 0)
                return 0
            
            confidence_scores = [get_ai_confidence_score(acc) for acc in accounts]
            avg_confidence = round(sum(confidence_scores) / len(accounts), 1) if accounts else 0
            
            summary_metrics = [
                ["Total Excel Rows", len(original_data)],
                ["Accounts Analyzed", len(accounts)],
                ["Average AI Confidence Score", avg_confidence],
                ["High Confidence (>80)", sum(1 for score in confidence_scores if score > 80)],
                ["Medium Confidence (50-80)", sum(1 for score in confidence_scores if 50 <= score <= 80)],
                ["Low Confidence (<50)", sum(1 for score in confidence_scores if score < 50)]
            ]
            
            for metric in summary_metrics:
                ws[f'A{current_row}'] = metric[0]
                ws[f'B{current_row}'] = metric[1]
                ws[f'A{current_row}'].font = Font(bold=True)
                ws[f'A{current_row}'].border = self.border
                ws[f'B{current_row}'].border = self.border
                current_row += 1
            
            current_row += 2
            
            # Add Analysis Table Header
            ws.merge_cells(f'A{current_row}:Z{current_row}')
            ws[f'A{current_row}'] = "📋 Original Data + AI Analysis"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="2C5AA0")
            ws[f'A{current_row}'].fill = PatternFill(start_color="F1EFEC", end_color="F1EFEC", fill_type="solid")
            current_row += 1
            
            # Create mapping of Account IDs to analysis results
            # FIX: Handle both 15-character and 18-character ID formats for matching
            # Issue: Excel input contains 15-character IDs, but Salesforce analysis returns 18-character IDs
            # Solution: Store both formats in the mapping to enable proper matching
            account_analysis_map = {}
            for acc in accounts:
                account_id = acc.get('Id', '')
                if account_id:
                    # Store with the 18-character ID (from analysis results)
                    account_analysis_map[account_id] = acc
                    # Also store with the 15-character ID for matching with original Excel data
                    if len(account_id) == 18:
                        account_analysis_map[account_id[:15]] = acc
                    elif len(account_id) == 15:
                        # If we have a 15-character ID, also store the 18-character version
                        account_analysis_map[self._convert_15_to_18_char_id(account_id)] = acc
            
            # Get original headers and add AI analysis columns
            if original_data:
                # Filter out any analysis-related fields from original data
                analysis_fields = {'AI_Assessment', 'Has_Shell', 'Customer_Consistency', 'Customer_Shell_Coherence', 'Address_Consistency', 'Shell_Account_Data'}
                original_headers = [header for header in original_data[0].keys() if header not in analysis_fields]
                ai_headers = ["AI Confidence Score", "AI Analysis"]
                all_headers = original_headers + ai_headers
                
                # Add headers
                for col, header in enumerate(all_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = self.header_font
                    if header in ai_headers:
                        cell.fill = PatternFill(start_color="00A3E0", end_color="00A3E0", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="0684BC", end_color="0684BC", fill_type="solid")
                    cell.alignment = self.center_alignment
                    cell.border = self.border
                current_row += 1
                
                # Add data rows
                for row_idx, row_data in enumerate(original_data, current_row):
                    # Get original data (filtered to exclude analysis fields)
                    row_values = []
                    for header in original_headers:
                        value = row_data.get(header, '')
                        
                        # Handle RecordType field - extract Name from dictionary
                        if header == 'RecordType' and isinstance(value, dict):
                            value = value.get('Name', 'N/A')
                        # Handle any other dictionary values
                        elif isinstance(value, dict):
                            value = str(value)  # Convert to string representation
                        
                        row_values.append(value)
                    
                    # Get AI analysis for this account
                    # FIX: Handle different column names for account IDs in original Excel data
                    # Issue: Original Excel data may have account IDs in user-selected column, not 'Id'
                    # Solution: Try multiple approaches to find the account ID
                    account_id = row_data.get('Id', '')
                    
                    # If 'Id' is not found, try to find the account ID column from excel_info
                    if not account_id and excel_info and 'account_id_column' in excel_info:
                        account_id_column = excel_info['account_id_column']
                        account_id = row_data.get(account_id_column, '')
                    
                    # If still no account_id found, try to find any field that looks like an account ID
                    if not account_id:
                        for key, value in row_data.items():
                            if isinstance(value, str) and len(value) in [15, 18] and value.startswith('001'):
                                account_id = value
                                break
                    
                    account_analysis = account_analysis_map.get(account_id, {})
                    
                    # If no match found, try to find by converting the ID
                    if not account_analysis and account_id:
                        if len(account_id) == 15:
                            # Try with 18-character version
                            converted_id = self._convert_15_to_18_char_id(account_id)
                            account_analysis = account_analysis_map.get(converted_id, {})
                        elif len(account_id) == 18:
                            # Try with 15-character version
                            account_analysis = account_analysis_map.get(account_id[:15], {})
                    
                    ai_assessment = account_analysis.get('AI_Assessment', {})
                    
                    # Handle AI assessment data safely - check if it's a string that needs parsing
                    if isinstance(ai_assessment, str):
                        try:
                            import json
                            ai_assessment = json.loads(ai_assessment)
                        except (json.JSONDecodeError, TypeError):
                            ai_assessment = {}
                    
                    confidence_score = ai_assessment.get('confidence_score', 0) if isinstance(ai_assessment, dict) else 0
                    explanation_bullets = ai_assessment.get('explanation_bullets', []) if isinstance(ai_assessment, dict) else []
                    
                    ai_values = [
                        f"{confidence_score}/100",
                        "\n".join([f"• {bullet}" for bullet in explanation_bullets]) if explanation_bullets else "No AI analysis available"
                    ]
                    
                    # Combine original and AI data
                    all_values = row_values + ai_values
                    
                    for col, value in enumerate(all_values, 1):
                        cell = ws.cell(row=row_idx, column=col, value=value)
                        cell.border = self.border
                
                # No frozen panes for Excel input export
                
                # Auto-adjust column widths
                for col in range(1, len(all_headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 20
            
            # Create file buffer
            file_buffer = io.BytesIO()
            wb.save(file_buffer)
            file_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"excel_analysis_{timestamp}.xlsx"
            
            return {
                'success': True,
                'file_buffer': file_buffer,
                'filename': filename
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f"Error creating Excel analysis export: {str(e)}"
            } 